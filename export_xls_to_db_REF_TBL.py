# 2020.08.05 Export data from xls file to db table.

import os
import openpyxl
import datetime

# mysql connector module
import mysql.connector

# declare mysql connection. Parameters: host-access address, user-ID, password, database-database name to access
dbconn = mysql.connector(host="localhost", user="root", passwd="xxxx", database="OOOOO")

# declare global variables
global exist_error
exist_error = False
global exit_now
exit_now = False

# Change the working directory
os.chdir('C:\\Users\\KKKKK\\Downloads\\DIR_NAME')

# function to match location for data clensing
def load_location_mapping():
    global exit_now
    global exist_error

    # Get the excel file
    cur_date = datetime.datetime.now()
    print( '>>> 1. Open an excel file(XXXXXXX.xlsx - ' + cur_date.strftime( '%Y/%m/%d %H:%M:%S' ) )
    # Get a cell's value not equation
    wb = openpyxl.load_workbook( 'XXXXXXX.xlsx', data_only=True)

    cur_date = datetime.datetime.now()
    print( '>>> 2. Open an excel sheet(SSSSSSS) - ' + cur_date.strftime( '%Y/%m/%d %H:%M:%S' ) )
    ref_map_sheet = wb['SSSSSSS']

    # get max clumns and rows
    ref_map_max_row = ref_map_sheet.max_row
    ref_map_max_col = ref_map_sheet.max_column

    try:
        # Get a cursor
        cursor = dbconn.cursor()
        query = "DELETE FROM TABLE_NM1 WHERE COL_NM1 = 'AAAA'"
        cursor.execute(query)
        print( '>>> 3. Delete all data existed')
    except Exception as e:
        exist_error = True
        dbconn.rollback()
        print(e.message)
        exit()

    print( '>>>> 4. Insert new data in the TABLE_NM1 table')
    
    # for statement start ==============================
    # 1st line is header and skipped
    for j in range( 2, ref_map_max_row+1):

        # array for data : No. of columns - 4
        ref_map_rec = [0 for x in range(5)]

        for i in range(0,4):
            ref_map_rec[i] = ref_map_sheet.cell(row=j, column=i+1).value

        # set the value for "Duplicate Key"
        ref_map_rec[4] = ref_map_rec[3]
        
        # if there is no specific index, there is no more data
        # Add the check logic because we cannot make sure that there is no blank record.
        if ref_map_rec[0] == None:
            print( '************************** case 1 The row number where blank record is ' + str(j))
            exit_now = True
        # Add the check logic because we cannot make sure that the cell type of the specific is string
        elif str( type( ref_map_rec[0])) == '<class \'string\'>':
            if len( ref_map_rec[0] < 1 ):
                print( '************************** case 2 The row number where blank record is ' + str(j))
                exit_now = True
        # Add the check logic because we cannot make sure that the cell type of the specific is number
        else:
            tmp_str = str( ref_map_rec[0] )
            if tmp_str == None or len( tmp_str ) < 1:
                print( '************************** case 3 The row number where blank record is ' + str(j))
                exit_now = True
        
        if exit_now:
            print( '************************** exit_now: ' + str(exit_now) )
            break

        try:
            # Get a cursor
            cursor = dbconn.cursor()
            query = """INSERT INTO TABLE_NM1 ( 
                                COL_NM1, 
                                COL_NM2, 
                                COL_NM3, 
                                COL_NM4 )
                              VALUES ( %s, %s, %s, %s )
                        ON DUPLICATE KEY UPDATE
                        COL_NM1 = %s """
            cursor.execute( query, ref_map_rec) 
        except Exception as e:
            exist_error = True
            dbconn.rollback()
            print( '************************** case 4 The row number where error occurs is ' + str(j))
            print( e.message)
            break

    cur_date = datetime.datetime.now()
    print( '>>>> 5. Finish to insert new data in the TALBE_NM1 table'  + cur_date.strftime( '%Y/%m/%d %H:%M:%S' ) + 'Total count: ' + str(j) )
    # for statement end ==============================

# function to cleanse the data in TABLE_NM1
def clenansing_data_TABLE_NM1():
    global exit_now
    global exist_error

    cur_date = datetime.datetime.now()
    print('>>>> 1. Cleansing data in TABLE_NM1 table'  + cur_date.strftime( '%Y/%m/%d %H:%M:%S' ))

    cur_date = datetime.datetime.now()

    try:
        # get a cursor
        cursor = dbconn.cursor()
        query = "SELECT COL0, COL1, COL2, COL3 FROM TABLE_NM5"
        cursor.execute(query)
        result = cursor.fetchall()
        ind = 0
        
        for rec in result:
            ind = ind + 1
            cursor = dbconn.cursor()

            if rec[3] == None:
                print( '*************** The row number where None occurs:' + str(ind) )
                if rec[1] == 'AAAA':
                    query1 = "UPDATE TABLE_NM1 SET COL_NM4 = '' WHERE COL_NM1 = '" + rec[0] + "' AND COL_NM2 = '" + rec[2] + "'"
                elif rec[1] == 'BBBB':
                    query1 = "UPDATE TABLE_NM1 SET COL_NM4 = '' WHERE COL_NM1 = '" + rec[0] + "' AND COL_NM2 = '" + rec[2] + "' AND COL_NM4 = 'CCCC'"
            # replace "\n" with "%"
            elif "\n" in rec[2]:
                print( '!!!!!!!!!!!!!!!! The row number where \'\\n\' occurs: ' + str(ind))
                tmp_str = rec[2]
                if rec[1] == 'AAAA':
                    query1 = "UPDATE TABLE_NM1 SET COL_NM4 = '" + rec[3] + "' WHERE COL_NM1 = '" + rec[0] + "' AND COL_NM2 = '" + tmp_str.replace( "\n", "%" ) + "'"
                elif rec[1] == 'BBBB':
                    query1 = "UPDATE TABLE_NM1 SET COL_NM4 = '" + rec[3] + "' WHERE COL_NM1 = '" + rec[0] +  "' AND COL_NM2 = '" + tmp_str.replace( "\n", "%" ) + "' AND COL_NM4 = 'CCCC'"
            else:
                if rec[1] == 'AAAA':
                    query1 = "UPDATE TABLE_NM1 SET COL_NM4 = '" + rec[3] + "' WHERE COL_NM1 = '" + rec[0] + "' AND COL_NM2 = '" + rec[2] + "'"
                elif rec[1] == 'BBBB':
                    query1 = "UPDATE TABLE_NM1 SET COL_NM4 = '" + rec[3] + "' WHERE COL_NM1 = '" + rec[0] +  "' AND COL_NM2 = '" + rec[2] + "' AND COL_NM4 = 'CCCC'"
            
            print( "*****************" + str(ind) + " - " + query1)
            cursor.execute(query1)

        cur_date = datetime.datetime.now()
        print('>>>> 2. Finish Cleansing data in TABLE_NM1 table'  + cur_date.strftime( '%Y/%m/%d %H:%M:%S' ))
    except Exception as e:
        exist_error = True
        dbconn.rollback()
        print(e.message)
        exit()
#================================================================
print ( 'Select the Number(1-2)')
print ( '1. Insert data into TABLE_NM1')
print ( '2. Cleanse data in TABLE_NM1')

sel_input = int( input() )

if sel_input == 1:
    load_location_mapping()
elif sel_input == 2:
    clenansing_data_TABLE_NM1()
else:
    print( 'You typed the wrong number!! (Press 1 or 2)') 
    exit()

if exist_error == False:
    dbconn.commit()