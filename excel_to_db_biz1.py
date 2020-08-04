#2020.07.24 엑셀파일에서 데이터 추출하기

import os
import openpyxl
#mysql connector를 사용하기 위한 모듈 선언
import mysql.connector

# mysql connection을 선언한다. 파라미터로는  host 접속주소, user는 ID, passwd 비밀번호, database는 접속할 Database이다
dbconn = mysql.connector.connect(host="localhost", user="usr", passwd="password", database="database_nm")

# Change the working directory
os.chdir('PATH')

# Get the xls file
wb = openpyxl.load_workbook('file_nm')
# Get a sheet in a xls file
biz_sheet = wb['Sheet_NM']

# Get max columns and rows
biz_max_row = biz_sheet.max_row
biz_max_col = biz_sheet.max_column

# for statement start ==>
for j in range(2, biz_max_row):

    # Initialize a variable for biz_answer
    biz_ans = [0 for x in range(38)]

    biz_ans[0] = 'SK'
    biz_ans[1] = biz_sheet.cell(row=j, column=37).value

    for i in range( 1, 36):
        biz_ans[i+1] = biz_sheet.cell(row=j, column=i).value

    # Get cursor
    cursor = dbconn.cursor()
    query = """INSERT INTO TBL_NM1 (
                        COL1,
                        COL2,
                        COL3,
                        COL4,
                        COL5,
                        COL6,
                        COL7,
                        COL8 )
                    VALUES( %s, %s, %s, %s, %s, %s, %s, %s ) """
    
    cursor.execute(query, biz_ans)

# for statement end =>

dbconn.commit()


